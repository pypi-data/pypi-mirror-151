#!/usr/bin/python3
from openpyxl import load_workbook
from docopt import docopt
import time
import json
import os

OPT = '''Usage:
two-tables [options] <conf>

-p, --pdf     Convert to PDF format (requires LibreOffice)
-h, --help    Print this help

two-tables is free software licensed under the GNU General
Public License, version 3 or later.
'''
DAYS = [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


def parse_date(date):
#  print('Read original date {}'.format(date))
  date = int(date)
  d = date % 100
  m = (date // 100) % 100
  y = (date // 10000)
  return y, m, d


def is_leap(year):
  if year % 4 == 0:
    if year % 100 == 0:
      return (year % 400 == 0)
    return True
  return False


def get_days(year, month):
  if month != 2:
    return DAYS[month]
  else:
    if is_leap(year):
      return DAYS[month] + 1
    else:
      return DAYS[month]


def main():
  args = docopt(OPT)
  conf_file_name = args['<conf>']
  convert_to_pdf = args['--pdf']
  conf_file = open(conf_file_name)
  conf = json.load(conf_file)
  print(conf)
  FILENAME = conf['file']
  cell = conf['modifyCell']
  for file in FILENAME:
    print(f'===Processing {file}===')
    print('Updating data...', end='')
    wb = load_workbook(filename=file+".xlsx")
    ws = wb.active
    y, m, d = parse_date(ws[cell].value)
  #   print("Read date {}{}{}".format(y, m, d))
    d = d + 7
    Y = 2000 + y
    if d > get_days(Y, m):
      d -= get_days(Y, m)
      m += 1
      if m > 12:
        m = 1
        y += 1
    sy, sm, sd = str(y), str(m), str(d)
    if d < 10:
      sd = "0" + sd
    if m < 10:
      sm = "0" + sm
    res = sy + sm + sd
    ws[cell] = res
  #  print('Write date {}'.format(res))
    wb.save(file+".xlsx")
    print('done')
    if convert_to_pdf:
      print('Converting spreadsheet to PDF format...', end='')
      os.system(f'libreoffice --headless --convert-to pdf \'{file}\'.xlsx')
      print('done')

if __name__ == '__main__':
  main()